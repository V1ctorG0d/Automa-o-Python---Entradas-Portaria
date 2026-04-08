import os
import pathlib
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook

class Model:
    #Busca o arquivo excel(.xlsx) modificado mais recentemente dentro de um diretório específico.
    def find_lastet_excel(dict):
        folder_path = pathlib.Path(dict)

        #Filtrar apenas aquivos com extensão .xlsx
        file = list(folder_path.glob("*.xlsx"))
        if not file:
           return None
        else:
           last_file = max(file, key=lambda f: f.stat().st_mtime)
           return last_file

    #Retornar o caminho completo de destino concatenando o diretório base com o nome do arquivo.
    def get_file_source(path ,source_file):
        source_path = os.path.join(path, source_file)
        return source_path
    
    #Retorna o caminho completo de destino concatenando o diretório base com o nome do arquivo.
    def get_file_destination(path ,destination_file):
        destination_path = os.path.join(path, destination_file)
        return destination_path
    
    #Lê o aequivo Excel e retorna um DataFrame contendo apenas as colunas específicadas 
    def excel_data_read(source_path):
        df_source = pd.read_excel(source_path, usecols=['Part Number', 'Supplier Name', 'Commodities'])
        return df_source
    
    #Leitura da Planilha PTP 
    def read_excel_ptp(source_path):
        # Abre o workbook via openpyxl apenas para localizar a linha do cabeçalho
        wb_ptp = load_workbook(source_path)
        ws_ptp = wb_ptp.active

        ptp_required = ["FILIAL", "COD ITEM", "COD FORNEC", "TP PED", "NOTA"]

        # Localiza a linha onde os campos obrigatórios estão presentes
        try:
            header_idx = Model.find_header_row(ws_ptp, ptp_required)
            pandas_header = header_idx -1
        except:
            pandas_header = 0
        finally:
            wb_ptp.close()

        # Lê os dados efetivamente usando Pandas
        df_ptp = pd.read_excel(source_path, header=pandas_header)
        df_ptp.columns = df_ptp.columns.str.strip().str.upper()

        # Mapeamento para renomear colunas do padrão nacional para internacional/projeto
        rename_map = {}
        if "FILIAL" in df_ptp.columns:
            rename_map["FILIAL"] = "PLANT"
        if "COD ITEM" in df_ptp.columns:
            rename_map["COD ITEM"] = "PART NUMBER"
        if "COD FORNEC" in df_ptp.columns:
            rename_map["COD FORNEC"] = "SUPPLIER CODE"
        if "TP PED" in df_ptp.columns:
            rename_map["TP PED"] = "STAGE"
        if "DT RECEB" in df_ptp.columns:
            rename_map["DT RECEB"] = "DATA"
        if "NOTA" in df_ptp.columns:
            rename_map["NOTA"] = "NOTA"

        df_ptp.rename(columns=rename_map, inplace=True)

        # Garante que as colunas existam e formata os textos
        for col in ["PLANT", "PART NUMBER", "SUPPLIER CODE", "STAGE","NOTA"]:
            if col not in df_ptp:
                df_ptp[col] = ""
            else:
                df_ptp[col] = df_ptp[col].astype(str).str.upper().str.strip()
        
        return df_ptp
    
    #Leitura da Planilha ATA **NÃO UTILIZADA** Duvidas sobre o funcionamento.
    def read_excel_ata(destination_path):
        wb = load_workbook(destination_path)
        ws = wb.active

        ata_required = ["PLANT (2)", "PART NUMBER", "NEW SUPPLIER CODE", "STAGE", "DATA ENTRADA DE AMOSTRA", "EFETIVAÇÃO DE PROJETO ENTRADA DE PORTARIA", "NOTAS L3", "NOTAS L4"]
        header_row = Model.find_header_row(ws, ata_required)
        header = {
            str(ws.cell(row=header_row, column=c).value).strip().upper(): c for c in range(1, ws.max_column + 1)
        }

        # Mapeia colunas da planilha Excel para nomes lógicos do sistema
        ata_map = {
            "PLANT (2)": "PLANT",
            "PART NUMBER": "PART NUMBER",
            "NEW SUPPLIER CODE": "SUPPLIER CODE",
            "STAGE": "STAGE",
            "DATA ENTRADA DE AMOSTRA": "DATAREC1",
            "EFETIVAÇÃO DE PROJETO ENTRADA DE PORTARIA": "DATAREC2",
            "NOTAS L3": "NOTAS_L3",
            "NOTAS L4": "NOTAS L4"
        }

        col_map = {}
        for col_excel, logical_name in ata_map.items():
            if col_excel.upper() in header:
                col_map[logical_name] = header[col_excel.upper()]
            else:
                raise Exception(f"Coluna '{col_excel}' não encontrada na ATA!")
            
        return col_map
        
    #Varre as primeiras 100 linhas da planilha para encontrar a linha que contém todos os nomes das colunas exigidas
    def find_header_row(ws, required_cols):
        for row in range(1, 100):
            values = [str(ws.cell(row=row, column=c).value).strip().upper() if ws.cell(row=row, column=c).value else "" for c in range(1, ws.max_column + 1)]
            if all(any(req in v for v in values) for req in required_cols):
                return row
        raise Exception("Cabeçalho da ATA não encontrado automaticamente!")
    
    #Cruzamento dos dados entre as planilhas e salvamento das datas em ATA
    def update_ata_with_ptp(destination_path, source_path):
        # Prepara os dados de origem (PTP)
        ptp= Model.read_excel_ptp(source_path)
        ptp.set_index(["PLANT", "PART NUMBER", "SUPPLIER CODE", "STAGE", "NOTA"], inplace=True)

        # Carrega a planilha de destino (ATA)
        wb = load_workbook(destination_path)
        ws = wb.active
        
        required_cols = ["PLANT (2)", "PART NUMBER", "NEW SUPPLIER CODE", "STAGE", "DATA ENTRADA DE AMOSTRA", "EFETIVAÇÃO DE PROJETO ENTRADA DE PORTARIA", "NOTAS L3", "NOTAS L4"]
        header_row = Model.find_header_row(ws, required_cols)
        
        # Cria um mapeamento de {NOME_COLUNA: ÍNDICE_DA_COLUNA}
        header = {
            str(ws.cell(row=header_row, column=c).value).strip().upper(): c for c in range(1, ws.max_column + 1)
        }

        # Mapeia colunas da planilha Excel para nomes lógicos do sistema
        ata_map = {
            "PLANT (2)": "PLANT",
            "PART NUMBER": "PART NUMBER",
            "NEW SUPPLIER CODE": "SUPPLIER CODE",
            "STAGE": "STAGE",
            "DATA ENTRADA DE AMOSTRA": "DATAREC1",
            "EFETIVAÇÃO DE PROJETO ENTRADA DE PORTARIA": "DATAREC2",
            "NOTAS L3": "NOTAS_L3",
            "NOTAS L4": "NOTAS L4"
        }

        col_map = {}
        for col_excel, logical_name in ata_map.items():
            if col_excel.upper() in header:
                col_map[logical_name] = header[col_excel.upper()]
            else:
                raise Exception(f"Coluna '{col_excel}' não encontrada na ATA!")
            
        col_plant = col_map["PLANT"]
        col_pn = col_map["PART NUMBER"]
        col_supplier = col_map["SUPPLIER CODE"]
        col_stage = col_map["STAGE"]
        col_rec1 = col_map["DATAREC1"]
        col_rec2 = col_map["DATAREC2"]
        col_notas_l3 = col_map["NOTAS_L3"]
        col_notas_l4 = col_map["NOTAS L4"]

        # Define as regras de validação de Stages de acordo com o tipo de pedido
        stage3_valid = ["t1", "t2"]
        stage4_valid = ["t3", "t4", "t5"]

        # Itera pelas linhas da ATA para buscar correspondência no PTP
        for row in range(header_row + 1, ws.max_row + 1):
            plant = str(ws.cell(row, col_plant).value).upper().strip()
            pn = str(ws.cell(row, col_pn).value).upper().strip()
            supplier = str(ws.cell(row, col_supplier).value).upper().strip()
            stage_ata = str(ws.cell(row, col_stage).value).upper().strip()

            # Pega o prefixo (Ex: "L3" de "L3 - Gate 3")
            stage_ata_prefix = stage_ata.split()[0]

            # Busca no DataFrame PTP usando os índices correspondentes
            try:    
                # Buscamos no DataFrame filtrando pelas 3 colunas
                # Usamos .xs para busca em multi-índice ou um filtro simples:
                res = ptp.xs((plant, pn, supplier), level=["PLANT", "PART NUMBER", "SUPPLIER CODE"], drop_level=False)
        
                for _, row_ptp in res.reset_index().iterrows():
                    stage_ptp = str(row_ptp["STAGE"]).strip().upper()
                    data_ptp = row_ptp["DATA"]
                    notas_ptp = row_ptp["NOTA"]

                    # Se o dado for um Timestamp do Pandas (NaT ou NaN), o Excel pode corromper
                    if pd.isna(data_ptp):
                        continue

                    # Se for L3 e o stage do PTP for válido, preenche DATAREC1 (se vazio)
                    if stage_ata_prefix == "L3" and stage_ptp in stage3_valid:
                        if ws.cell(row, col_rec1).value is None:
                            ws.cell(row, col_rec1).value = data_ptp
                        # Insere a nota na coluna L3 se estiver vazia
                        if ws.cell(row, col_notas_l3).value is None:
                            ws.cell(row, col_notas_l3).value = notas_ptp

                    # Se for L4 e o stage do PTP for válido, preenche DATAREC2 (se vazio)
                    if stage_ata_prefix == "L4" and stage_ptp in stage4_valid:
                        if ws.cell(row, col_rec2).value is None:
                            ws.cell(row, col_rec2).value = data_ptp
                        # Insere a nota na coluna L4 se estiver vazia
                        if ws.cell(row, col_notas_l4).value is None:
                            ws.cell(row, col_notas_l4).value = notas_ptp

            except KeyError:
                # Se não encontrar a chave no PTP, pula para a próxima linha
                continue

        wb.save(destination_path)
        print(f"✔ Planilha {destination_path} atualizada com sucesso!")

    #Carrega um arquivo existente e adiciona um novo cabeçalho de colunas.
    def load_file(destination_path):
        book = load_workbook(destination_path)
        sheet = book.active
        sheet.append(['Part Number', 'Supplier Name', 'Commodities'])
        return sheet
    
    #Converte um DataFrame do Pandas em uma lista de listas (formato aceito pelo openpyxl).
    def convert_df_list(df_source):
        insert_data = df_source.values.tolist()
        return insert_data

    #Itera sobre uma lista de dados e adiciona cada linha ao final da planilha.
    def insert_data(insert_data, sheet):
        for line in insert_data:
            sheet.append(line)

    #Salva o objeto Workbook no caminho especificado.
    def save_file(book, destination_path):
        book.save(destination_path)
        print(f"Data was successfully inserted into{destination_path}")